"""회원 명단 임포트 (T-007) — 구글시트 다운로드 파일(CSV/XLSX)을 업서트한다.

운영 전제(결정 1): 구글시트가 직원 편집용 원본이고 이 임포터로 주기 동기화한다.
따라서 반복 실행 안전(멱등)이 최우선이다:
- 매칭: email → (name+organization) → 신규. 실명단에 이메일 결측·중복이 있어
  email은 unique가 아니다.
- 빈 셀은 기존 값을 지우지 않는다. 삭제는 자동으로 하지 않는다(사고 방지).
- dry_run=True면 DB에 쓰지 않고 리포트만 낸다.
"""

import csv
import io
import re
import secrets
from typing import Any

from sqlmodel import Session, select

from app.lib.logger import get_logger
from app.models.member import Member
from app.models.member_program import MemberProgram

logger = get_logger("member_import")

EXPECTED_HEADERS = {"성명", "이메일", "소분류", "구분", "휴대폰"}
EMPTY_VALUES = {"", "-", "—", "X", "x", "None", "nan"}


# ---------------------------------------------------------------- cleaners


def _norm(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return None if text in EMPTY_VALUES else text


def _norm_phone(value: Any) -> str | None:
    if not value:
        return None
    digits = re.sub(r"[^\d\-+]", "", str(value))
    return digits or None


def _first_email(value: Any) -> str | None:
    """셀에 이메일 여러 개(줄바꿈·세미콜론)면 첫 번째만 취한다."""
    if not value:
        return None
    text = str(value).replace("\n", ",").replace(";", ",")
    parts = [p.strip() for p in text.split(",") if "@" in p]
    return parts[0] if parts else None


def parse_row(row: dict[str, Any]) -> dict[str, Any] | None:
    """시트 한글 헤더 → Member 필드. 성명 없으면 스킵(None)."""
    name = _norm(row.get("성명"))
    if not name:
        return None
    notes = _norm(row.get("비고"))
    payment = notes if notes and ("납입" in notes or "이후" in notes) else None
    return {
        "name": name,
        "email": _first_email(row.get("이메일")),
        "phone": _norm_phone(row.get("휴대폰")),
        "cohort": _norm(row.get("소분류")),
        "category": _norm(row.get("구분")),
        "subcategory": _norm(row.get("세부구분")),
        "position": _norm(row.get("직위")),
        "organization": _norm(row.get("소속")),
        "location": _norm(row.get("소재지")),
        "division": _norm(row.get("세부소속")),
        "business_area": _norm(
            row.get("연구분야(대학소속) / 사업분야(기업소속)") or row.get("사업분야")
        ),
        "membership_status": _norm(row.get("* 총동문회"))
        or _norm(row.get("회원 여부 (2025.04.01. 기준)")),
        "membership_type": _norm(row.get("유료회원여부 (26.6.18.)")),
        "payment_history": payment,
        "benefit_pct": _norm(row.get("* 월드푸드테크협의회"))
        or _norm(row.get("혜택 적용 여부 (수강료 감면)")),
        "council_label": _norm(row.get("소속 (협의회 표기 기준)")),
        "notes": notes,
    }


# ---------------------------------------------------------------- file readers


def _find_header_row(rows: list[list[str]], max_check: int = 15) -> int:
    """시트 위쪽 제목·설명 행을 건너뛰고 실제 헤더 행을 찾는다."""
    best, best_score = 0, 0
    for i, row in enumerate(rows[:max_check]):
        score = sum(1 for cell in row if cell and cell.strip() in EXPECTED_HEADERS)
        if score > best_score:
            best, best_score = i, score
    return best


def _rows_to_dicts(all_rows: list[list[str]]) -> tuple[list[dict[str, Any]], int]:
    if not all_rows:
        return [], 0
    header_idx = _find_header_row(all_rows)
    headers = [h.strip() for h in all_rows[header_idx]]
    out = []
    for raw in all_rows[header_idx + 1 :]:
        if not any(raw):
            continue
        out.append(dict(zip(headers, raw, strict=False)))
    return out, header_idx + 1


def read_csv(content: bytes) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    text = None
    encoding = ""
    for enc in ("utf-8-sig", "utf-8", "cp949", "euc-kr"):
        try:
            text = content.decode(enc)
            encoding = enc
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("CSV 인코딩을 해석할 수 없습니다 (utf-8/cp949/euc-kr 시도)")
    all_rows = [[str(v).strip() if v else "" for v in row] for row in csv.reader(io.StringIO(text))]
    rows, header_row = _rows_to_dicts(all_rows)
    return rows, {"format": "csv", "encoding": encoding, "header_row": header_row}


def read_xlsx(
    content: bytes, sheet: str | None = None
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    if sheet is not None:
        if sheet not in wb.sheetnames:
            raise ValueError(f"시트 없음: {sheet} (있는 시트: {', '.join(wb.sheetnames[:10])})")
        ws = wb[sheet]
    else:
        ws = wb.active
    assert ws is not None
    all_rows = [
        ["" if v is None else str(v).strip() for v in row] for row in ws.iter_rows(values_only=True)
    ]
    rows, header_row = _rows_to_dicts(all_rows)
    return rows, {"format": "xlsx", "sheet": ws.title, "header_row": header_row}


def read_file(
    content: bytes, filename: str, sheet: str | None = None
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    if filename.lower().endswith((".xlsx", ".xlsm")):
        return read_xlsx(content, sheet)
    return read_csv(content)


# ---------------------------------------------------------------- upsert


class _MemberIndex:
    """행마다 DB 왕복을 피하기 위한 메모리 매칭 인덱스.

    원격 DB(Supabase) 대상 4,000행 임포트가 행당 SELECT 2회로는 10분+ 걸려서,
    시작 시 전체를 한 번 로드하고 이후는 메모리에서 매칭한다.
    """

    def __init__(self, session: Session) -> None:
        self.by_email: dict[str, Member] = {}
        self.by_name_org: dict[tuple[str, str | None], Member] = {}
        for m in session.exec(select(Member)).all():
            self.register(m)

    def register(self, member: Member) -> None:
        if member.email and member.email not in self.by_email:
            self.by_email[member.email] = member
        key = (member.name, member.organization)
        if key not in self.by_name_org:
            self.by_name_org[key] = member

    def find(self, data: dict[str, Any]) -> Member | None:
        email = data.get("email")
        if email and email in self.by_email:
            return self.by_email[email]
        candidate = self.by_name_org.get((data["name"], data.get("organization")))
        if candidate is None:
            return None
        # 동명이인 보호: 양쪽 다 이메일이 있는데 서로 다르면 같은 이름·소속이라도 다른 사람으로
        # 취급. (이메일이 한쪽만 있으면 같은 사람의 결측 보완으로 보고 병합)
        if email and candidate.email and candidate.email != email:
            return None
        return candidate


def import_members(
    session: Session,
    content: bytes,
    filename: str,
    *,
    program: str | None = None,
    sheet: str | None = None,
    dry_run: bool = False,
) -> dict[str, Any]:
    """파일 1개를 업서트하고 리포트를 반환한다. dry_run이면 롤백."""
    rows, file_info = read_file(content, filename, sheet)
    report: dict[str, Any] = {
        **file_info,
        "rows": len(rows),
        "created": 0,
        "updated": 0,
        "skipped": 0,
        "linked_program": 0,
        "errors": [],
        "dry_run": dry_run,
    }
    index = _MemberIndex(session)
    pending_links: list[tuple[Member, str | None]] = []
    for i, row in enumerate(rows, start=file_info["header_row"] + 1):
        try:
            data = parse_row(row)
            if data is None:
                report["skipped"] += 1
                continue
            cohort = data.pop("cohort")
            member = index.find(data)
            if member is None:
                member = Member(**data, unsubscribe_token=secrets.token_urlsafe(24))
                session.add(member)
                report["created"] += 1
            else:
                # 빈 셀(None)은 기존 값을 지우지 않는다 — 부분 채움 시트도 안전
                for key, value in data.items():
                    if value is not None:
                        setattr(member, key, value)
                if not member.unsubscribe_token:
                    member.unsubscribe_token = secrets.token_urlsafe(24)
                session.add(member)
                report["updated"] += 1
            index.register(member)
            if program:
                pending_links.append((member, cohort))
        except Exception as exc:  # 행 단위 격리 — 한 행 오류가 전체를 막지 않는다
            report["errors"].append({"row": i, "error": str(exc)})

    if program and pending_links:
        session.flush()  # 신규 회원 id 일괄 확보
        existing_links = {
            (link.member_id, link.program): link
            for link in session.exec(
                select(MemberProgram).where(MemberProgram.program == program)
            ).all()
        }
        for member, cohort in pending_links:
            assert member.id is not None
            link = existing_links.get((member.id, program))
            if link is None:
                link = MemberProgram(member_id=member.id, program=program, cohort=cohort)
                session.add(link)
                existing_links[(member.id, program)] = link
                report["linked_program"] += 1
            elif cohort and not link.cohort:
                link.cohort = cohort
                session.add(link)
    if dry_run:
        session.rollback()
    else:
        session.commit()
    logger.info(
        f"member import: rows={report['rows']} created={report['created']} "
        f"updated={report['updated']} skipped={report['skipped']} "
        f"errors={len(report['errors'])} dry_run={dry_run}"
    )
    return report
